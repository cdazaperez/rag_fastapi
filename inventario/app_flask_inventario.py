# app_flask_inventario.py
from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response
from werkzeug.utils import secure_filename
import mysql.connector
import os
from decimal import Decimal
from datetime import datetime
import json
from io import BytesIO
from xhtml2pdf import pisa
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage

app = Flask(__name__)
app.secret_key = 'clave_secreta_segura'
UPLOAD_FOLDER = 'static/uploads'
LOGO_FOLDER = 'static/logos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['LOGO_FOLDER'] = LOGO_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOGO_FOLDER, exist_ok=True)

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'inventario_user'),
    'password': os.getenv('DB_PASSWORD', 'Sk1llN3t.2025'),
    'database': os.getenv('DB_NAME', 'inventario_uniformes')
}

# Formateo de moneda en COP para usar en las plantillas
@app.template_filter('cop')
def formato_cop(value):
    try:
        value = float(value)
        formatted = f"{value:,.2f}"
        formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"COP {formatted}"
    except (TypeError, ValueError):
        return value

def get_db():
    return mysql.connector.connect(**DB_CONFIG)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_config():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM configuracion LIMIT 1")
    config = cursor.fetchone()
    if not config:
        cursor.execute(
            "INSERT INTO configuracion(nombre_empresa) VALUES ('Mi Empresa')")
        conn.commit()
        cursor.execute("SELECT * FROM configuracion LIMIT 1")
        config = cursor.fetchone()
    conn.close()
    return config

@app.context_processor
def inject_config():
    return {'configuracion': get_config()}

def send_email(to_email, subject, body, pdf_bytes):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = 'noreply@example.com'
    msg['To'] = to_email
    msg.set_content(body)
    if pdf_bytes:
        msg.add_attachment(pdf_bytes, maintype='application', subtype='pdf', filename='recibo.pdf')
    try:
        with smtplib.SMTP('localhost') as smtp:
            smtp.send_message(msg)
    except Exception as e:
        print(f"Error enviando email: {e}")

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM usuarios WHERE username=%s AND password=%s", (username, password))
        user = cursor.fetchone()
        conn.close()
        if user:
            session['user'] = username
            session['role'] = user['role']
            return redirect(url_for('dashboard'))
        flash("Credenciales inválidas", "error")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM productos")
    productos = cursor.fetchall()
    conn.close()
    return render_template('dashboard.html', productos=productos)

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre FROM colegios")
    colegios = cursor.fetchall()

    if request.method == 'POST':
        data = (
            request.form['referencia'],
            request.form['nombre'],
            request.form['colegio'],
            request.form['genero'],
            request.form['talla'],
            int(request.form['cantidad']),
            float(request.form['precio']),
            id
        )
        cursor.execute(
            "UPDATE productos SET referencia=%s, nombre=%s, colegio=%s, genero=%s, talla=%s, cantidad=%s, precio=%s WHERE id=%s",
            data,
        )
        conn.commit()
        conn.close()
        flash("Producto actualizado")
        return redirect(url_for('dashboard'))
    cursor.execute("SELECT * FROM productos WHERE id=%s", (id,))
    producto = cursor.fetchone()
    conn.close()
    return render_template('editar.html', producto=producto, colegios=colegios)

@app.route('/venta/<int:id>', methods=['POST'])
def venta(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    cantidad = int(request.form['cantidad'])
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT cantidad FROM productos WHERE id=%s", (id,))
    stock = cursor.fetchone()[0]
    if stock >= cantidad:
        cursor.execute("UPDATE productos SET cantidad=cantidad-%s WHERE id=%s", (cantidad, id))
        cursor.execute("INSERT INTO ventas(producto_id, usuario, cantidad, fecha) VALUES (%s, %s, %s, %s)", 
                       (id, session['user'], cantidad, datetime.now()))
        conn.commit()
        flash("Venta registrada")
    else:
        flash("Stock insuficiente", "error")
    conn.close()
    return redirect(url_for('dashboard'))

@app.route('/usuarios')
def gestionar_usuarios():
    if session.get('role') != 'admin':
        flash("Acceso denegado", "error")
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM usuarios")
    usuarios = cursor.fetchall()
    conn.close()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuarios/crear', methods=['POST'])
def crear_usuario():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))

    username = request.form['username']
    password = request.form['password']
    email = request.form.get('email')
    role = request.form['role'].strip().lower()

    if role not in ("admin", "user"):
        flash("Rol inválido", "error")
        return redirect(url_for('gestionar_usuarios'))

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO usuarios(username, password, email, role) VALUES (%s, %s, %s, %s)",
            (username, password, email, role),
        )
        conn.commit()
        flash("Usuario creado")
    except mysql.connector.Error as e:
        flash(f"Error al crear usuario: {e}", "error")
    finally:
        conn.close()

    return redirect(url_for('gestionar_usuarios'))

@app.route('/usuarios/borrar/<int:id>', methods=['POST'])
def eliminar_usuario(id):
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM usuarios WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Usuario eliminado")
    return redirect(url_for('gestionar_usuarios'))


@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM configuracion LIMIT 1")
    config = cursor.fetchone()

    if request.method == 'POST':
        nombre = request.form.get('nombre_empresa')
        direccion = request.form.get('direccion')
        telefono = request.form.get('telefono')
        email = request.form.get('email')
        color = request.form.get('color_primario')
        logo_file = request.files.get('logo')
        logo_path = config.get('logo') if config else None
        if logo_file and allowed_file(logo_file.filename):
            filename = secure_filename(logo_file.filename)
            logo_path = os.path.join(app.config['LOGO_FOLDER'], filename)
            logo_file.save(logo_path)

        if config:
            cursor.execute(
                """UPDATE configuracion SET nombre_empresa=%s, direccion=%s, telefono=%s, email=%s, logo=%s, color_primario=%s WHERE id=%s""",
                (nombre, direccion, telefono, email, logo_path, color, config['id'])
            )
        else:
            cursor.execute(
                """INSERT INTO configuracion(nombre_empresa, direccion, telefono, email, logo, color_primario)
                VALUES (%s, %s, %s, %s, %s, %s)""",
                (nombre, direccion, telefono, email, logo_path, color)
            )
        conn.commit()
        flash('Configuración actualizada')
        cursor.execute("SELECT * FROM configuracion LIMIT 1")
        config = cursor.fetchone()

    conn.close()
    return render_template('configuracion.html', config=config)

@app.route('/eliminar/<int:id>', methods=['POST'])
def eliminar_producto(id):
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ventas WHERE producto_id=%s", (id,))
    cursor.execute("DELETE FROM productos WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Producto eliminado")
    return redirect(url_for('dashboard'))

@app.route('/ventas')
def ver_ventas():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT v.id, p.nombre, p.colegio, v.usuario, v.cantidad, p.precio,
               (v.cantidad * p.precio) AS valor, v.fecha
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        ORDER BY p.colegio, v.fecha DESC
        """
    )
    ventas = cursor.fetchall()
    conn.close()
    return render_template('ventas.html', ventas=ventas)

@app.route('/reportes')
def reporte_ventas():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT DATE(fecha) AS dia, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY dia ORDER BY dia DESC
    """)
    diarios = cursor.fetchall()
    cursor.execute("""
        SELECT YEAR(fecha) AS year, WEEK(fecha) AS semana, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY year, semana ORDER BY year DESC, semana DESC
    """)
    semanal = cursor.fetchall()
    cursor.execute("""
        SELECT DATE_FORMAT(fecha, '%Y-%m') AS mes, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY mes ORDER BY mes DESC
    """)
    mensual = cursor.fetchall()
    cursor.execute("""
        SELECT usuario, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY usuario ORDER BY usuario
    """)
    por_usuario = cursor.fetchall()
    cursor.execute("""
        SELECT p.colegio, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY p.colegio ORDER BY p.colegio
    """)
    por_colegio = cursor.fetchall()
    conn.close()
    return render_template('reportes.html', diarios=diarios, semanal=semanal, mensual=mensual,
                           por_usuario=por_usuario, por_colegio=por_colegio)


@app.route('/reportes/excel')
def exportar_excel_reportes():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT DATE(fecha) AS dia, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY dia ORDER BY dia DESC
        """
    )
    diarios = cursor.fetchall()
    cursor.execute(
        """
        SELECT YEAR(fecha) AS year, WEEK(fecha) AS semana, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY year, semana ORDER BY year DESC, semana DESC
        """
    )
    semanal = cursor.fetchall()
    cursor.execute(
        """
        SELECT DATE_FORMAT(fecha, '%Y-%m') AS mes, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY mes ORDER BY mes DESC
        """
    )
    mensual = cursor.fetchall()
    cursor.execute(
        """
        SELECT usuario, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY usuario ORDER BY usuario
        """
    )
    por_usuario = cursor.fetchall()
    cursor.execute(
        """
        SELECT p.colegio, SUM(v.cantidad * p.precio) AS total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        GROUP BY p.colegio ORDER BY p.colegio
        """
    )
    por_colegio = cursor.fetchall()
    conn.close()

    wb = Workbook()

    ws = wb.active
    ws.title = 'Diario'
    ws.append(['Día', 'Total'])
    for r in diarios:
        ws.append([r['dia'], float(r['total'])])

    ws = wb.create_sheet('Semanal')
    ws.append(['Año', 'Semana', 'Total'])
    for r in semanal:
        ws.append([r['year'], r['semana'], float(r['total'])])

    ws = wb.create_sheet('Mensual')
    ws.append(['Mes', 'Total'])
    for r in mensual:
        ws.append([r['mes'], float(r['total'])])

    ws = wb.create_sheet('Por Usuario')
    ws.append(['Usuario', 'Total'])
    for r in por_usuario:
        ws.append([r['usuario'], float(r['total'])])

    ws = wb.create_sheet('Por Colegio')
    ws.append(['Colegio', 'Total'])
    for r in por_colegio:
        ws.append([r['colegio'], float(r['total'])])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reportes.xlsx'
    return response

@app.route('/agregar', methods=['GET', 'POST'])
def agregar_producto():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre FROM colegios")
    colegios = cursor.fetchall()

    if request.method == 'POST':
        referencia = request.form['referencia']
        nombre = request.form['nombre']
        colegio = request.form['colegio']
        genero = request.form['genero']
        talla = request.form['talla']
        cantidad = int(request.form['cantidad'])
        precio = float(request.form['precio'])

        cursor.execute(
            """
            INSERT INTO productos (referencia, nombre, colegio, genero, talla, cantidad, precio)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (referencia, nombre, colegio, genero, talla, cantidad, precio),
        )
        conn.commit()
        conn.close()
        flash("Producto agregado correctamente")
        return redirect(url_for('dashboard'))

    conn.close()
    return render_template("agregar_producto.html", colegios=colegios)


@app.route('/venta', methods=['GET', 'POST'])
def nueva_venta():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM productos")
    productos = cursor.fetchall()
    cursor.execute("SELECT DISTINCT colegio FROM productos")
    colegios = [row['colegio'] for row in cursor.fetchall()]
    cursor.execute("SELECT id, nombre, apellido, email FROM clientes")
    clientes = cursor.fetchall()

    if request.method == 'POST':
        items = json.loads(request.form['items']) if request.form.get('items') else []
        aplicar_iva = request.form.get('aplicar_iva') == 'on'
        cliente_id = request.form.get('cliente_id')
        if not cliente_id:
            flash('Debe seleccionar un cliente', 'error')
            conn.close()
            return redirect(url_for('nueva_venta'))

        cursor.execute("SELECT * FROM clientes WHERE id=%s", (cliente_id,))
        cliente = cursor.fetchone()
        if not cliente:
            flash('Cliente no encontrado', 'error')
            conn.close()
            return redirect(url_for('nueva_venta'))

        ventas_detalle = []
        total = Decimal('0.00')

        for item in items:
            producto_id = int(item['id'])
            cantidad = int(item['cantidad'])
            cursor.execute(
                "SELECT nombre, colegio, cantidad, precio FROM productos WHERE id=%s",
                (producto_id,))
            prod = cursor.fetchone()
            if not prod or prod['cantidad'] < cantidad:
                flash("Stock insuficiente o producto no encontrado", "error")
                conn.close()
                return redirect(url_for('nueva_venta'))

            precio = Decimal(str(prod['precio']))
            subtotal = precio * cantidad
            ventas_detalle.append({
                'nombre': prod['nombre'],
                'colegio': prod['colegio'],
                'cantidad': cantidad,
                'precio': float(precio),
                'subtotal': float(subtotal)
            })
            total += subtotal

            cursor.execute(
                "UPDATE productos SET cantidad = cantidad - %s WHERE id = %s",
                (cantidad, producto_id))
            cursor.execute(
                "INSERT INTO ventas(producto_id, usuario, cliente_id, cantidad, fecha) VALUES (%s, %s, %s, %s, %s)",
                (producto_id, session['user'], cliente_id, cantidad, datetime.now()))

        iva = total * Decimal('0.19') if aplicar_iva else Decimal('0.00')
        total_final = total + iva

        session['last_sale'] = {
            'ventas': ventas_detalle,
            'total': float(total),
            'aplicar_iva': aplicar_iva,
            'iva': float(iva),
            'total_final': float(total_final),
            'cliente': cliente
        }

        conn.commit()
        conn.close()
        return render_template(
            "recibo_venta.html",
            ventas=ventas_detalle,
            total=total,
            aplicar_iva=aplicar_iva,
            iva=iva,
            total_final=total_final,
            cliente=cliente
        )

    conn.close()
    return render_template("nueva_venta.html", productos=productos, colegios=colegios, clientes=clientes)


@app.route('/recibo/pdf')
def recibo_pdf():
    if 'last_sale' not in session:
        flash('No hay recibo disponible', 'error')
        return redirect(url_for('dashboard'))

    data = session['last_sale']
    html = render_template('recibo_venta_pdf.html', **data)
    pdf_stream = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_stream)
    if pisa_status.err:
        flash('Error al generar PDF', 'error')
        return redirect(url_for('dashboard'))
    pdf_stream.seek(0)
    response = make_response(pdf_stream.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=recibo.pdf'
    return response

@app.route('/recibo/enviar')
def enviar_recibo():
    if 'last_sale' not in session:
        flash('No hay recibo disponible', 'error')
        return redirect(url_for('dashboard'))

    data = session['last_sale']
    html = render_template('recibo_venta_pdf.html', **data)
    pdf_stream = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_stream)
    if pisa_status.err:
        flash('Error al generar PDF', 'error')
        return redirect(url_for('dashboard'))
    pdf_stream.seek(0)
    cliente = data.get('cliente')
    if cliente and cliente.get('email'):
        send_email(
            cliente['email'],
            'Recibo de compra',
            'Adjuntamos su recibo de compra.',
            pdf_stream.read()
        )
        flash('Recibo enviado al cliente', 'success')
    else:
        flash('El cliente no tiene email registrado', 'error')
    return redirect(url_for('dashboard'))

# --- CRUD Colegios ---

@app.route('/colegios')
def listar_colegios():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM colegios")
    colegios = cursor.fetchall()
    conn.close()
    return render_template('colegios.html', colegios=colegios)


@app.route('/colegios/crear', methods=['POST'])
def crear_colegio():
    if 'user' not in session:
        return redirect(url_for('login'))
    nombre = request.form['nombre']
    direccion = request.form['direccion']
    telefono = request.form['telefono']
    contacto = request.form['contacto']
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO colegios(nombre, direccion, telefono, contacto) VALUES (%s, %s, %s, %s)",
        (nombre, direccion, telefono, contacto))
    conn.commit()
    conn.close()
    flash('Colegio creado correctamente')
    return redirect(url_for('listar_colegios'))


@app.route('/colegios/editar/<int:id>', methods=['GET', 'POST'])
def editar_colegio(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    if request.method == 'POST':
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        contacto = request.form['contacto']
        cursor.execute(
            "UPDATE colegios SET nombre=%s, direccion=%s, telefono=%s, contacto=%s WHERE id=%s",
            (nombre, direccion, telefono, contacto, id))
        conn.commit()
        conn.close()
        flash('Colegio actualizado')
        return redirect(url_for('listar_colegios'))
    cursor.execute("SELECT * FROM colegios WHERE id=%s", (id,))
    colegio = cursor.fetchone()
    conn.close()
    return render_template('editar_colegio.html', colegio=colegio)


@app.route('/colegios/borrar/<int:id>', methods=['POST'])
def eliminar_colegio(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM colegios WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash('Colegio eliminado')
    return redirect(url_for('listar_colegios'))


# --- CRUD Clientes ---

@app.route('/clientes')
def listar_clientes():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()
    conn.close()
    return render_template('clientes.html', clientes=clientes)


@app.route('/clientes/crear', methods=['POST'])
def crear_cliente():
    if 'user' not in session:
        return redirect(url_for('login'))
    nombre = request.form['nombre']
    apellido = request.form['apellido']
    cedula = request.form['cedula']
    telefono = request.form.get('telefono')
    email = request.form.get('email')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO clientes(nombre, apellido, cedula, telefono, email) VALUES (%s, %s, %s, %s, %s)",
        (nombre, apellido, cedula, telefono, email))
    conn.commit()
    conn.close()
    flash('Cliente creado correctamente')
    return redirect(url_for('listar_clientes'))


@app.route('/clientes/editar/<int:id>', methods=['GET', 'POST'])
def editar_cliente(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        cedula = request.form['cedula']
        telefono = request.form.get('telefono')
        email = request.form.get('email')
        cursor.execute(
            """UPDATE clientes SET nombre=%s, apellido=%s, cedula=%s, telefono=%s, email=%s WHERE id=%s""",
            (nombre, apellido, cedula, telefono, email, id))
        conn.commit()
        conn.close()
        flash('Cliente actualizado')
        return redirect(url_for('listar_clientes'))
    cursor.execute("SELECT * FROM clientes WHERE id=%s", (id,))
    cliente = cursor.fetchone()
    conn.close()
    return render_template('editar_cliente.html', cliente=cliente)


@app.route('/clientes/borrar/<int:id>', methods=['POST'])
def eliminar_cliente(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM clientes WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash('Cliente eliminado')
    return redirect(url_for('listar_clientes'))


# --- Gastos ---

@app.route('/gastos')
def listar_gastos():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, item, descripcion, cantidad, valor_unitario, fecha FROM gastos ORDER BY fecha DESC")
    gastos = cursor.fetchall()
    cursor.execute("SELECT DATE(fecha) AS dia, SUM(cantidad * valor_unitario) AS total FROM gastos GROUP BY dia ORDER BY dia DESC")
    diarios = cursor.fetchall()
    cursor.execute("SELECT YEAR(fecha) AS year, WEEK(fecha) AS semana, SUM(cantidad * valor_unitario) AS total FROM gastos GROUP BY year, semana ORDER BY year DESC, semana DESC")
    semanal = cursor.fetchall()
    cursor.execute("SELECT DATE_FORMAT(fecha, '%Y-%m') AS mes, SUM(cantidad * valor_unitario) AS total FROM gastos GROUP BY mes ORDER BY mes DESC")
    mensual = cursor.fetchall()
    conn.close()
    return render_template('gastos.html', gastos=gastos, diarios=diarios, semanal=semanal, mensual=mensual)


@app.route('/gastos/crear', methods=['POST'])
def crear_gasto():
    if 'user' not in session:
        return redirect(url_for('login'))
    item = request.form['item']
    descripcion = request.form.get('descripcion')
    cantidad = int(request.form['cantidad'])
    valor_unitario = float(request.form['valor_unitario'])
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO gastos(item, descripcion, cantidad, valor_unitario, fecha) VALUES (%s, %s, %s, %s, %s)",
        (item, descripcion, cantidad, valor_unitario, datetime.now()))
    conn.commit()
    conn.close()
    flash('Gasto registrado correctamente')
    return redirect(url_for('listar_gastos'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8082, debug=True)
